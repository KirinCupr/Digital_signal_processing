import os
import numpy as np
from scipy.signal.windows import hann

import openpyxl

# ---
# def fft(y, fs) - функция, которая принимает на вход: массив y - дискретная последовательность сигналов; fs - частоту дискретизации
# Возвращает два массива
# Использование:
# xf, yf = fft(y, fs)
# ---

def fft(y, fs):
    xf=np.fft.rfftfreq(n=y.size,d=1/fs)
    yf=np.abs(np.fft.rfft(hann(y.size)*y))*2/y.size
    return xf, yf


# ---
# def xlsx_to_txt(path) - функция для создания каталога с txt-файлами данных из файла xlsx, полученного в системе Адамант(?)
# На каждую страницу создается отдельный txt-файл
# Использование: 
# xlsx_to_txt('test.xlsx')
# ---

def xlsx_to_txt(path):
    wb_obj = openpyxl.load_workbook(path)
    sheets = wb_obj.sheetnames

    path = path[:-5]  + '_folder'

    while True:
        try:
            os.mkdir(path)
            for sheet_obj in sheets:
                filepath = os.path.join(path, sheet_obj + '.txt')
                with open(filepath, 'w+') as file:
                    for j in range(1, wb_obj[sheet_obj].max_column + 1):
                        for i in range(13, wb_obj[sheet_obj].max_row + 1):
                            cell_obj = wb_obj[sheet_obj].cell(row = i, column = j)
                            file.write(str(cell_obj.value) + '\n')
            print('Папка ', path, ' успешно создана!')
            break
            
        except OSError as error:
            print(error)
            path = input("Ой! Папка с таким уже есть. Пожалуйста, введите свое имя \n>> ")


# ---
# def all_xlsx_to_txt() - функция для создания каталогов с txt-файлами данных из ВСЕХ файлов xlsx, полученных в системе Адамант(?)
# Для каждого файла создается отдельный каталог с именем "имя_файла_folder"
# использование: 
# all_xlsx_to_txt()
# ---
                    
def all_xlsx_to_txt():
    files = os.listdir()
    for file in files:
        if file[-5:] == '.xlsx':
            xlsx_to_txt(file)
